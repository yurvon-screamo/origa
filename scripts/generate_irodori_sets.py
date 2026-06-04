"""Generate Irodori well_known_set JSON files from wordlist_all.xlsx."""

from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = PROJECT_ROOT / "wordlist_all.xlsx"
OUTPUT_BASE = PROJECT_ROOT / "cdn" / "well_known_set"
META_PATH = OUTPUT_BASE / "well_known_types_meta.json"

COL_HEADING = 3
COL_READING = 4
COL_ENGLISH = 7
COL_SECTION = 10
COL_LESSON = 11
COL_REF_MARKER = 16

SECTION_MAP: dict[str, dict] = {
    "入門": {
        "dir_name": "irodori_nyuumon",
        "level": "N5",
        "book_ru": "入門",
        "book_en": "Nyuumon",
    },
    "初級1": {
        "dir_name": "irodori_shokyuu1",
        "level": "N4",
        "book_ru": "初級1",
        "book_en": "Shokyuu 1",
    },
    "初級2": {
        "dir_name": "irodori_shokyuu2",
        "level": "N4",
        "book_ru": "初級2",
        "book_en": "Shokyuu 2",
    },
}


def clean_word(raw: str) -> str:
    if not raw:
        return ""
    text = str(raw).strip()
    if not text:
        return ""
    full_bracket_prefix = re.compile(r"^（[^）]+）")
    full_bracket_prefix_matches = full_bracket_prefix.match(text)
    if full_bracket_prefix_matches and len(full_bracket_prefix_matches.group()) == len(text):
        return ""
    if text.startswith("（"):
        return full_bracket_prefix.sub("", text).strip()
    text = re.sub(r"（([^）]+)）", r"\1", text)
    return text


def build_groups(wb) -> dict[tuple[str, int], set[str]]:
    ws = wb.active
    groups: dict[tuple[str, int], set[str]] = defaultdict(set)
    skipped_ref = 0
    skipped_empty = 0
    skipped_cleaned = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < COL_REF_MARKER:
            continue
        ref_marker = row[COL_REF_MARKER - 1]
        if str(ref_marker).strip() == "参":
            skipped_ref += 1
            continue
        heading = row[COL_HEADING - 1]
        section = row[COL_SECTION - 1]
        lesson = row[COL_LESSON - 1]
        if not heading or not section or not lesson:
            skipped_empty += 1
            continue
        section_str = str(section).strip()
        lesson_int = int(lesson)
        if section_str not in SECTION_MAP:
            continue
        cleaned = clean_word(heading)
        if not cleaned:
            skipped_cleaned += 1
            continue
        groups[(section_str, lesson_int)].add(cleaned)
    print(f"Filtered: {skipped_ref} reference words skipped")
    print(f"Filtered: {skipped_empty} rows skipped (missing data)")
    print(f"Filtered: {skipped_cleaned} words cleaned to empty")
    return groups


def write_json(file_path: Path, section: str, lesson: int, words: list[str]) -> None:
    meta = SECTION_MAP[section]
    word_count = len(words)
    data = {
        "content": {
            "Russian": {
                "title": f"Irodori {meta['book_ru']} Урок {lesson}",
                "description": f"Слова урока {lesson} ({word_count} {'слово' if word_count == 1 else 'слов'})",
            },
            "English": {
                "title": f"Irodori {meta['book_en']} Lesson {lesson}",
                "description": f"Lesson {lesson} words ({word_count} {'word' if word_count == 1 else 'words'})",
            },
        },
        "level": meta["level"],
        "words": words,
    }
    file_path.parent.mkdir(parents=True, exist_ok=True)
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")


def update_meta() -> None:
    with open(META_PATH, "r", encoding="utf-8") as f:
        meta = json.load(f)
    existing_ids = {t["id"] for t in meta["types"]}
    if "Irodori" not in existing_ids:
        meta["types"].append(
            {"id": "Irodori", "label_ru": "Irodori", "label_en": "Irodori"}
        )
        meta["types"].sort(key=lambda t: t["id"])
        with open(META_PATH, "w", encoding="utf-8") as f:
            json.dump(meta, f, ensure_ascii=False, indent=2)
            f.write("\n")
        print("Updated well_known_types_meta.json with Irodori entry")
    else:
        print("well_known_types_meta.json already contains Irodori entry")


def main() -> None:
    if not EXCEL_PATH.exists():
        print(f"Error: {EXCEL_PATH} not found")
        return
    print(f"Loading {EXCEL_PATH} ...")
    wb = load_workbook(EXCEL_PATH, read_only=True)
    groups = build_groups(wb)
    wb.close()
    print(f"\nFound {len(groups)} lesson groups")
    total_files = 0
    total_words = 0
    summary: list[str] = []
    for (section, lesson), words in sorted(groups.items()):
        meta = SECTION_MAP[section]
        dir_name = meta["dir_name"]
        file_name = f"{dir_name}_{lesson:02d}.json"
        file_path = OUTPUT_BASE / dir_name / file_name
        sorted_words = sorted(words)
        write_json(file_path, section, lesson, sorted_words)
        total_files += 1
        total_words += len(sorted_words)
        summary.append(f"  {dir_name}/{file_name}: {len(sorted_words)} words")
    for line in summary:
        print(line)
    print(f"\nTotal: {total_files} files, {total_words} words")
    update_meta()
    print("Done.")


if __name__ == "__main__":
    main()
